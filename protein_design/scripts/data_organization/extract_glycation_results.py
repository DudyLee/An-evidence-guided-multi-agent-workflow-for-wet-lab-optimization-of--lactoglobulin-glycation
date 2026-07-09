import csv
import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
RAW_WORKBOOK_DIR = os.environ.get("GLYCATION_RAW_WORKBOOK_DIR")
SOURCE_DIR = Path(RAW_WORKBOOK_DIR) if RAW_WORKBOOK_DIR else None
OUT_DIR = ROOT / "protein_design" / "tables" / "glycation_results_organized"

TIMECOURSE_FILE = "2026.3.19-\u5355\u7cd6\u4e8c\u7cd6-\u7cd6\u57fa\u5316-\u6570\u636e\u6574\u7406.xlsx"
ELISA_0615_FILE = "ELISA-2026.6.15.xlsx"

FORMAL_24_CSV = ROOT / "protein_design" / "tables" / "wetlab_runsheet" / "formal_24_run_result_mapping.csv"
FIG3_SOURCE_CSV = ROOT / "protein_design" / "figures" / "wetlab_results_v2" / "Figure_3_v2_source_data.csv"
FIG3C_SOURCE_CSV = ROOT / "protein_design" / "figures" / "wetlab_results_v2" / "Figure_3c_v2_source_data.csv"


DONOR_MAP = {
    "\u6838\u7cd6": "ribose",
    "\u6728\u7cd6": "xylose",
    "\u963f\u62c9\u4f2f\u7cd6": "arabinose",
    "\u8461\u8404\u7cd6": "glucose",
    "\u679c\u7cd6": "fructose",
    "\u534a\u4e73\u7cd6": "galactose",
    "\u4e73\u7cd6": "lactose",
    "\u9ea6\u82bd\u7cd6": "maltose",
    "\u7ea4\u7ef4\u4e8c\u7cd6": "cellobiose",
    "\u7518\u9732\u7cd6": "mannose",
}


def safe_float(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except Exception:
        return None


def round_or_none(value, ndigits=4):
    return None if value is None else round(float(value), ndigits)


def read_csv_dicts(path):
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fieldnames=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        seen = []
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.append(key)
        fieldnames = seen
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def workbook_role(file_name):
    if file_name == TIMECOURSE_FILE:
        return "long donor-time ultrasound/no-ultrasound panel and ribose 4 h validation point"
    if file_name == ELISA_0615_FILE:
        return "supplemental ELISA records completing R09-R12, R19 and no-sugar controls"
    if "filled_run_mapping" in file_name:
        return "historical run-mapping workbook; superseded by current 24-run matrix"
    if "RPBG-4h" in file_name:
        return "4 h hexose bridge and no-sugar controls for Fig. 3a/3b"
    if "2026.1.21" in file_name:
        return "2 h dry glycation benchmark/pentose support used in the run matrix"
    if "2025.12.12" in file_name:
        return "arabinose-focused supporting workbook"
    return "raw wet-lab glycation/ELISA workbook"


def scan_workbooks():
    if SOURCE_DIR is None:
        raise RuntimeError(
            "This legacy extraction script requires the original raw wet-lab "
            "workbooks. Set GLYCATION_RAW_WORKBOOK_DIR before running it."
        )
    files = sorted(
        [p for p in SOURCE_DIR.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm"}],
        key=lambda p: p.name.lower(),
    )
    file_inventory = []
    sheet_summary = []
    raw_cells = []

    for path in files:
        stat = path.stat()
        wb_values = load_workbook(path, read_only=True, data_only=True)
        wb_formulas = load_workbook(path, read_only=True, data_only=False)
        sheet_names = wb_values.sheetnames
        file_inventory.append(
            {
                "file_name": path.name,
                "file_path": str(path),
                "size_bytes": stat.st_size,
                "last_modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "sheet_count": len(sheet_names),
                "sheet_names": "; ".join(sheet_names),
                "interpreted_role": workbook_role(path.name),
            }
        )

        for ws_values in wb_values.worksheets:
            ws_formulas = wb_formulas[ws_values.title]
            nonempty_rows = 0
            nonempty_cells = 0
            keyword_counts = {
                "3h": 0,
                "4h": 0,
                "6h": 0,
                "9h": 0,
                "12h": 0,
                "18h": 0,
                "24h": 0,
                "US_or_ultrasound": 0,
                "ribose": 0,
                "glucose": 0,
                "arabinose": 0,
            }
            for row_idx, row in enumerate(ws_values.iter_rows(values_only=True), start=1):
                row_has_value = False
                formula_row = next(
                    ws_formulas.iter_rows(
                        min_row=row_idx,
                        max_row=row_idx,
                        values_only=False,
                    )
                )
                for col_idx, value in enumerate(row, start=1):
                    if value is None:
                        continue
                    row_has_value = True
                    nonempty_cells += 1
                    text = str(value)
                    low = text.lower()
                    if len(raw_cells) < 50000:
                        formula_cell = formula_row[col_idx - 1]
                        raw_cells.append(
                            {
                                "file_name": path.name,
                                "sheet_name": ws_values.title,
                                "row": row_idx,
                                "column": col_idx,
                                "cell": formula_cell.coordinate,
                                "value": value,
                                "formula_or_display": formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None,
                            }
                        )
                    for k in ["3h", "4h", "6h", "9h", "12h", "18h", "24h"]:
                        if k in low or k.replace("h", " h") in low:
                            keyword_counts[k] += 1
                    if "us" in low or "\u8d85\u58f0" in low:
                        keyword_counts["US_or_ultrasound"] += 1
                    if "ribose" in low or "\u6838\u7cd6" in low:
                        keyword_counts["ribose"] += 1
                    if "glucose" in low or "\u8461\u8404\u7cd6" in low:
                        keyword_counts["glucose"] += 1
                    if "arabinose" in low or "\u963f\u62c9\u4f2f\u7cd6" in low:
                        keyword_counts["arabinose"] += 1
                if row_has_value:
                    nonempty_rows += 1
            sheet_summary.append(
                {
                    "file_name": path.name,
                    "sheet_name": ws_values.title,
                    "max_row": ws_values.max_row,
                    "max_column": ws_values.max_column,
                    "nonempty_rows": nonempty_rows,
                    "nonempty_cells": nonempty_cells,
                    **keyword_counts,
                }
            )
    return file_inventory, sheet_summary, raw_cells


def extract_timecourse():
    if SOURCE_DIR is None:
        raise RuntimeError(
            "extract_timecourse requires GLYCATION_RAW_WORKBOOK_DIR to point to "
            "the original raw wet-lab workbooks."
        )
    path = SOURCE_DIR / TIMECOURSE_FILE
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    native_baseline = safe_float(rows[1][1])
    records = []
    validation_records = []

    for excel_row, row in enumerate(rows[2:], start=3):
        label = row[0]
        if not isinstance(label, str):
            continue
        match = re.match(r"(.+)-(\d+)h", label)
        if not match:
            continue
        donor_cn = match.group(1)
        time_h = int(match.group(2))
        donor = DONOR_MAP.get(donor_cn, donor_cn)
        for ultrasound, mean_idx, sd_idx in [
            ("no US", 1, 2),
            ("+US", 4, 5),
        ]:
            mean = safe_float(row[mean_idx])
            sd = safe_float(row[sd_idx])
            if mean is None:
                continue
            reduction = (1 - mean / native_baseline) * 100 if native_baseline else None
            error_pct = (sd / native_baseline) * 100 if sd is not None and native_baseline else None
            records.append(
                {
                    "dataset": "long donor-time panel",
                    "source_file": path.name,
                    "source_sheet": "Sheet1",
                    "source_row": excel_row,
                    "donor_cn": donor_cn,
                    "donor": donor,
                    "time_h": time_h,
                    "temperature_C": 55,
                    "temperature_note": "Source sheet does not print temperature; treated as the same dry-state framework used in the manuscript after user confirmation.",
                    "ultrasound": ultrasound,
                    "mode": "dry",
                    "aw": 0.79,
                    "native_baseline": round_or_none(native_baseline, 6),
                    "binding_capacity_mean": round_or_none(mean, 6),
                    "binding_capacity_sd": round_or_none(sd, 6),
                    "reduction_pct_vs_native": round_or_none(reduction, 4),
                    "error_pct_vs_native": round_or_none(error_pct, 4),
                    "use_in_manuscript": "candidate Supplementary Fig. 1 / Supplementary Data 4",
                }
            )

    # The right-side cells in row 3 store the independent ribose 4 h validation point.
    validation_label = rows[2][10]
    validation_mean = safe_float(rows[2][11])
    validation_reduction = safe_float(rows[2][12])
    validation_error = safe_float(rows[2][13])
    if validation_mean is not None and validation_reduction is not None:
        validation_records.append(
            {
                "dataset": "final ribose validation",
                "source_file": path.name,
                "source_sheet": "Sheet1",
                "source_cells": "K3:N3",
                "source_label": validation_label,
                "donor": "ribose",
                "time_h": 4,
                "temperature_C": 60,
                "ultrasound": "+US",
                "mode": "dry",
                "aw": 0.79,
                "binding_capacity_mean": round_or_none(validation_mean, 6),
                "reduction_pct_vs_native": round_or_none(validation_reduction * 100, 4),
                "error_pct": round_or_none(validation_error * 100, 4),
                "use_in_manuscript": "Fig. 3c validation point; candidate Supplementary Fig. 1 annotation",
                "note": "Standalone post-recommendation validation point for the final ribose condition.",
            }
        )
    return records, validation_records


def extract_elisa_0615():
    if SOURCE_DIR is None:
        raise RuntimeError(
            "extract_elisa_0615 requires GLYCATION_RAW_WORKBOOK_DIR to point to "
            "the original raw wet-lab workbooks."
        )
    path = SOURCE_DIR / ELISA_0615_FILE
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    native = safe_float(rows[1][2])
    mapping = {
        "N": ("none", None, None, "native reference", None, "native untreated baseline"),
        "N-U-2h": ("none", 55, 2, "+US", "R21", "no-sugar process control"),
        "N-U-3h": ("none", 55, 3, "+US", "R22", "no-sugar process control; Chinese text repeats 2 h but sample code and user confirmation specify 3 h"),
        "\u6838\u7cd6-U-2h": ("ribose", 55, 2, "+US", "R19", "ribose pentose branch"),
        "\u4e73\u7cd6-4h": ("lactose", 55, 4, "no US", "R09", "hexose no-ultrasound sentinel"),
        "\u8461\u8404\u7cd6-4h": ("glucose", 55, 4, "no US", "R10", "hexose no-ultrasound sentinel"),
        "\u534a\u4e73\u7cd6-4h": ("galactose", 55, 4, "no US", "R11", "hexose no-ultrasound sentinel"),
        "\u7518\u9732\u7cd6-4h": ("mannose", 55, 4, "no US", "R12", "hexose no-ultrasound sentinel"),
    }
    records = []
    for excel_row, row in enumerate(rows[1:], start=2):
        condition_text, sample_code, mean, sd = row[:4]
        mean = safe_float(mean)
        sd = safe_float(sd)
        donor, temp, time_h, ultrasound, run_id, role = mapping.get(sample_code, (None, None, None, None, None, None))
        reduction = (1 - mean / native) * 100 if mean is not None and native else None
        error_pct = (sd / native) * 100 if sd is not None and native else None
        records.append(
            {
                "source_file": path.name,
                "source_sheet": "Sheet1",
                "source_row": excel_row,
                "condition_text": condition_text,
                "sample_code": sample_code,
                "mapped_run_id": run_id,
                "mapped_role": role,
                "donor": donor,
                "temperature_C": temp,
                "time_h": time_h,
                "ultrasound": ultrasound,
                "native_baseline": round_or_none(native, 6),
                "binding_capacity_mean": round_or_none(mean, 6),
                "binding_capacity_sd": round_or_none(sd, 6),
                "reduction_pct_vs_native": round_or_none(reduction, 4),
                "error_pct_vs_native": round_or_none(error_pct, 4),
                "use_in_manuscript": "completes 24-run matrix for Fig. 3a/3b and Supplementary Tables 3-5",
            }
        )
    return records


def build_figure_data_map(file_inventory, formal_24, timecourse_records, validation_records):
    source_files = {row["file_name"]: row["interpreted_role"] for row in file_inventory}
    figure_map = []
    usage = [
        ("Main Fig. 3a", "R01-R12 matched 4 h hexose bridge and no-ultrasound sentinels", "formal_24_run_result_mapping.csv / Figure_3_v2_source_data.csv"),
        ("Main Fig. 3b", "R13-R24 short-window anchors, pentose branch and no-sugar process controls", "formal_24_run_result_mapping.csv / Figure_3_v2_source_data.csv"),
        ("Main Fig. 3c left", "glucose 55 C +US marginal-gain time course", "Figure_3c_v2_source_data.csv and 2026.3.19 workbook"),
        ("Main Fig. 3c right", "agent estimate versus ribose 60 C 4 h +US wet-lab validation", "Figure_3c_v2_source_data.csv and 2026.3.19 workbook"),
        ("Candidate Supplementary Fig. 1", "complete donor-time panel with no-US and +US arms", TIMECOURSE_FILE),
        ("Supplementary Tables 3-5", "complete R01-R24 condition/result audit and wet-lab values", "formal_24_run_result_mapping.csv / Supplementary_Data_3_24_run_result_matrix.xlsx"),
        ("Supplementary Data 4", "raw and processed glycation/ELISA data package", "all Excel workbooks in glycation result folder"),
    ]
    for item, content, source in usage:
        figure_map.append({"item": item, "content": content, "source": source, "note": ""})

    source_count = {}
    for row in formal_24:
        source = row.get("source_file") or ""
        source_count[source] = source_count.get(source, 0) + 1
    for source, count in sorted(source_count.items()):
        figure_map.append(
            {
                "item": "24-run source contribution",
                "content": f"{count} planned run(s) mapped to this source file",
                "source": source,
                "note": source_files.get(source, ""),
            }
        )
    return figure_map


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    file_inventory, sheet_summary, raw_cells = scan_workbooks()
    timecourse_records, validation_records = extract_timecourse()
    elisa_records = extract_elisa_0615()
    formal_24 = read_csv_dicts(FORMAL_24_CSV)
    fig3_source = read_csv_dicts(FIG3_SOURCE_CSV)
    fig3c_source = read_csv_dicts(FIG3C_SOURCE_CSV)
    figure_map = build_figure_data_map(file_inventory, formal_24, timecourse_records, validation_records)

    datasets = {
        "file_inventory": file_inventory,
        "sheet_summary": sheet_summary,
        "raw_cell_index": raw_cells,
        "protected_24_run_matrix": formal_24,
        "fig3_source_data": fig3_source,
        "fig3c_source_data": fig3c_source,
        "long_timecourse": timecourse_records,
        "elisa_2026_0615": elisa_records,
        "ribose_validation": validation_records,
        "figure_data_map": figure_map,
    }

    for name, rows in datasets.items():
        write_csv(OUT_DIR / f"{name}.csv", rows)

    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_dir": str(SOURCE_DIR) if SOURCE_DIR else None,
        "out_dir": str(OUT_DIR),
        "datasets": datasets,
        "summary": {
            "source_workbook_count": len(file_inventory),
            "sheet_count": len(sheet_summary),
            "raw_cell_index_rows": len(raw_cells),
            "protected_24_run_rows": len(formal_24),
            "long_timecourse_rows": len(timecourse_records),
            "elisa_2026_0615_rows": len(elisa_records),
            "ribose_validation_rows": len(validation_records),
        },
    }
    (OUT_DIR / "glycation_results_organized_payload.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))
    print(str(OUT_DIR))


if __name__ == "__main__":
    main()
